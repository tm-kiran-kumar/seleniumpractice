import openpyxl

path = r'/Users/kiran/Developer/AutomationProjects/seleniumpractice/data/New_Spreadsheet.xlsx'


def get_excel_data():
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    final_list = []
    for row in sheet.iter_rows(min_row=2, max_col=0, values_only=True):
        username, password = row
        final_list.append(row)
    return final_list

data = get_excel_data()


# def get_data_from_excel(path):
#     # Load the workbook and select the active sheet
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook.active
#
#     final_list = []
#
#     # Iterate through rows, starting from row 2 to skip headers
#     # sheet.max_row gets the total number of rows with data
#     for r in range(2, sheet.max_row + 1):
#         # Read data from Column 1 (A) and Column 2 (B)
#         username = sheet.cell(row=r, column=1).value
#         password = sheet.cell(row=r, column=2).value
#
#         # Add tuple to list: ('student', 'Password123')
#         final_list.append((username, password))
#
#     return final_list