from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

# Create a new object
# wb = Workbook()
# # Get active worksheet (created by default)
# ws = wb.active
# ws.title = 'Sample1'
# wb.save('New_Spreadsheet.xlsx')

# Load Workbook
# filepath = 'New_Spreadsheet.xlsx'
# wb = load_workbook(filepath)
# ws = wb.active
#
# ws['A1'] = 'Username'
# # ws['B1'] = 'Password'
# ws['A2'], ws['B2'] = 'admin0', 'pass0'
# ws['A3'], ws['B3'] = 'admin1', 'pass1'

# wb.save(filepath)

# To Read rows in for loop
# for row in ws.iter_rows(min_row=2, max_col=0, values_only=True):
#     #print(row)
#     username, password = row

# # Using function
# def auth_user_pass(data):
#     for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
#         # print(row)
#         username, password = row
#     return username, password
#
# username, password = auth_user_pass(ws)
# print(f'Excel credentials- {username} and {password}')
#

# For multiple rows and columns
# user_list = []
# def get_details(ws):
#     for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
#         user_list.append(row)
#
#     return user_list
#
# res = get_details(ws)
# print(res)
